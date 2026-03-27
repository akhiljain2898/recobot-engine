"""
excel_generator.py — easemyreco Excel Report Generator

7-Sheet structure:
  Sheet 1 — Summary
  Sheet 2 — Matched (Exact)
  Sheet 3 — Matched (Variation)
  Sheet 4 — A Not in B
  Sheet 5 — B Not in A
  Sheet 6 — Payment Mismatches
  Sheet 7 — No Invoice Number

Changes (v2):
  [1] Sheets 4 and 5 renamed to standard "A Not in B" / "B Not in A"
  [2] Reason tag column added to Sheets 4, 5, and 7
  [3] Sheet 7 added for blank invoice number entries from both parties
      with a Party column to distinguish A vs B
  [4] Summary sheet — party A/B defined at top, reason breakdown table added below
  [5] No-invoice entries excluded from Sheet 4/5 counts — live on Sheet 7 only
"""

from io import BytesIO
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
# Style constants
# ─────────────────────────────────────────────
TEAL         = "00838F"
GREEN        = "2E7D32"
ORANGE       = "E65100"
RED          = "B71C1C"
PURPLE       = "6A1B9A"
GREY         = "ECEFF1"
LIGHT_TEAL   = "E0F7FA"
LIGHT_GREEN  = "E8F5E9"
LIGHT_ORANGE = "FFF3E0"
LIGHT_RED    = "FFEBEE"
LIGHT_PURPLE = "F3E5F5"
WHITE        = "FFFFFF"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size)

def _border() -> Border:
    thin = Side(style="thin", color="BDBDBD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _apply_header_row(ws, headers: list, row: int = 1,
                      bg: str = TEAL, fg: str = WHITE):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font      = _font(bold=True, color=fg, size=10)
        cell.fill      = _fill(bg)
        cell.alignment = _center()
        cell.border    = _border()


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)


def _fmt_amount(v) -> str:
    try:
        return f"₹{float(v):,.2f}"
    except (TypeError, ValueError):
        return "₹0.00"


def _get(e, key, default=""):
    return e.get(key) or default


def _entry_cols(e) -> tuple:
    return (
        _get(e, "invoice_number_raw"),
        _get(e, "date"),
        _get(e, "voucher_type"),
        _get(e, "amount", 0.0),
    )


# ─────────────────────────────────────────────
# Sheet 1 — Summary
# ─────────────────────────────────────────────

def _build_summary(ws, results: dict, party_a: str, party_b: str, period: dict):
    ws.title = "Summary"

    inv = results["invoices"]
    pay = results["payments"]

    # Title
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value     = "easemyreco — Reconciliation Report"
    title.font      = Font(bold=True, size=14, color=WHITE)
    title.fill      = _fill(TEAL)
    title.alignment = _center()

    # FIX [4]: Party A/B defined clearly at top
    info = [
        ("Party A", party_a),
        ("Party B", party_b),
        ("Period From", period.get("from", "")),
        ("Period To",   period.get("to",   "")),
        ("Report Date", datetime.now().strftime("%d-%b-%Y")),
    ]
    for r_offset, (label, value) in enumerate(info, start=2):
        lbl_cell = ws.cell(row=r_offset, column=1, value=label)
        lbl_cell.font = _font(bold=True)
        ws.cell(row=r_offset, column=2, value=value)

    row = len(info) + 3

    # Invoice Reconciliation block
    ws.cell(row=row, column=1, value="INVOICE RECONCILIATION").font = _font(bold=True, size=11)
    row += 1
    summary_inv = [
        ("Total invoices — A",           inv["total_a"]),
        ("Total invoices — B",           inv["total_b"]),
        ("Matched (Exact)",              inv["matched_exact"]),
        ("Matched (Variation ≤1%)",      inv["matched_variation"]),
        ("In A, not in B — Count",       inv["a_not_in_b_count"]),
        ("In A, not in B — Value",       _fmt_amount(inv["a_not_in_b_value"])),
        ("In B, not in A — Count",       inv["b_not_in_a_count"]),
        ("In B, not in A — Value",       _fmt_amount(inv["b_not_in_a_value"])),
    ]
    for label, value in summary_inv:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    # Payment Reconciliation block
    ws.cell(row=row, column=1, value="PAYMENT RECONCILIATION").font = _font(bold=True, size=11)
    row += 1
    summary_pay = [
        ("Total payments — A",           pay["total_a"]),
        ("Total payments — B",           pay["total_b"]),
        ("Matched",                      pay["matched"]),
        ("In A, not in B — Count",       pay["a_not_in_b_count"]),
        ("In A, not in B — Value",       _fmt_amount(pay["a_not_in_b_value"])),
        ("In B, not in A — Count",       pay["b_not_in_a_count"]),
        ("In B, not in A — Value",       _fmt_amount(pay["b_not_in_a_value"])),
    ]
    for label, value in summary_pay:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 2

    # FIX [4]: Reason breakdown table below existing summary
    ws.cell(row=row, column=1, value="UNMATCHED ENTRY BREAKDOWN").font = _font(bold=True, size=11)
    row += 1
    _apply_header_row(ws, ["Reason", "Count", "Value"], row=row, bg=TEAL)
    row += 1

    breakdown = [
        (
            "No invoice number found in A",
            inv.get("no_invoice_count_a", 0),
            _fmt_amount(inv.get("no_invoice_value_a", 0.0)),
        ),
        (
            "No invoice number found in B",
            inv.get("no_invoice_count_b", 0),
            _fmt_amount(inv.get("no_invoice_value_b", 0.0)),
        ),
        (
            "Invoice found in A — no match in B",
            inv["a_not_in_b_count"],
            _fmt_amount(inv["a_not_in_b_value"]),
        ),
        (
            "Invoice found in B — no match in A",
            inv["b_not_in_a_count"],
            _fmt_amount(inv["b_not_in_a_value"]),
        ),
    ]

    # Only show fuzzy_skipped row if there are any
    fuzzy_entries = (
        [e for e in inv.get("_a_not_in_b", [])
         if e.get("unmatched_reason", "").startswith("Match attempt incomplete")] +
        [e for e in inv.get("_b_not_in_a", [])
         if e.get("unmatched_reason", "").startswith("Match attempt incomplete")]
    )
    if fuzzy_entries:
        fuzzy_val = sum(float(e.get("amount", 0.0)) for e in fuzzy_entries)
        breakdown.append((
            "Match attempt incomplete — dataset too large",
            len(fuzzy_entries),
            _fmt_amount(fuzzy_val),
        ))

    for reason, count, value in breakdown:
        ws.cell(row=row, column=1, value=reason).border = _border()
        ws.cell(row=row, column=2, value=count).border  = _border()
        ws.cell(row=row, column=3, value=value).border  = _border()
        row += 1

    _auto_width(ws)


# ─────────────────────────────────────────────
# Sheet 2 — Matched (Exact)
# ─────────────────────────────────────────────

def _build_matched_exact(ws, matched: list, party_a: str, party_b: str):
    ws.title = "Matched (Exact)"
    headers = [
        "Invoice No (A)", "Date (A)", "Type (A)", "Amount (A)",
        "Invoice No (B)", "Date (B)", "Type (B)", "Amount (B)",
    ]
    _apply_header_row(ws, headers, bg=GREEN)
    for row_idx, m in enumerate(matched, start=2):
        ea = m["entry_a"]
        eb = m["entry_b"]
        inv_a, date_a, vtype_a, amt_a = _entry_cols(ea)
        inv_b, date_b, vtype_b, amt_b = _entry_cols(eb)
        vals = [inv_a, date_a, vtype_a, _fmt_amount(amt_a),
                inv_b, date_b, vtype_b, _fmt_amount(amt_b)]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.fill      = _fill(LIGHT_GREEN)
            cell.alignment = _left()
            cell.border    = _border()
    _auto_width(ws)


# ─────────────────────────────────────────────
# Sheet 3 — Matched (Variation)
# ─────────────────────────────────────────────

def _build_matched_variation(ws, matched: list, party_a: str, party_b: str):
    ws.title = "Matched (Variation)"
    headers = [
        "Invoice No (A)", "Date (A)", "Amount (A)",
        "Invoice No (B)", "Date (B)", "Amount (B)",
        "Difference (A−B)", "Variation %",
    ]
    _apply_header_row(ws, headers, bg=ORANGE)
    for row_idx, m in enumerate(matched, start=2):
        ea = m["entry_a"]
        eb = m["entry_b"]
        inv_a, date_a, _, amt_a = _entry_cols(ea)
        inv_b, date_b, _, amt_b = _entry_cols(eb)
        diff = m.get("difference", round(float(amt_a) - float(amt_b), 2))
        pct  = round(abs(diff) / max(abs(float(amt_a)), abs(float(amt_b)), 0.01) * 100, 2) \
               if amt_a or amt_b else 0
        vals = [
            inv_a, date_a, _fmt_amount(amt_a),
            inv_b, date_b, _fmt_amount(amt_b),
            _fmt_amount(diff), f"{pct}%",
        ]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.fill      = _fill(LIGHT_ORANGE)
            cell.alignment = _left()
            cell.border    = _border()
    _auto_width(ws)


# ─────────────────────────────────────────────
# Sheets 4 & 5 — A Not in B / B Not in A
# FIX [1]: standardised names
# FIX [2]: reason tag column added
# ─────────────────────────────────────────────

def _build_unmatched_single(ws, entries: list, sheet_title: str):
    ws.title = sheet_title
    headers = ["Invoice No", "Date", "Voucher Type", "Amount", "Reason"]
    _apply_header_row(ws, headers, bg=RED)
    for row_idx, e in enumerate(entries, start=2):
        inv, date, vtype, amt = _entry_cols(e)
        reason = e.get("unmatched_reason") or ""
        vals = [inv, date, vtype, _fmt_amount(amt), reason]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.fill      = _fill(LIGHT_RED)
            cell.alignment = _left()
            cell.border    = _border()
    _auto_width(ws)


# ─────────────────────────────────────────────
# Sheet 6 — Payment Mismatches
# ─────────────────────────────────────────────

def _build_payment_mismatches(ws, results_pay: dict, party_a: str, party_b: str):
    ws.title = "Payment Mismatches"

    ws.cell(row=1, column=1, value="MATCHED PAYMENTS").font = _font(bold=True, size=11)
    _apply_header_row(ws, ["Date (A)", "Amount (A)", "Date (B)", "Amount (B)"], row=2, bg=GREEN)

    row = 3
    for m in results_pay.get("_matched", []):
        ea = m["entry_a"]
        eb = m["entry_b"]
        vals = [ea.get("date"), _fmt_amount(ea.get("amount", 0)),
                eb.get("date"), _fmt_amount(eb.get("amount", 0))]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=col_idx, value=v)
            cell.fill   = _fill(LIGHT_GREEN)
            cell.border = _border()
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="IN A, NOT IN B").font = _font(bold=True, size=11)
    row += 1
    _apply_header_row(ws, ["Date", "Amount", "Voucher Type"], row=row, bg=RED)
    row += 1
    for e in results_pay.get("_a_not_in_b", []):
        vals = [e.get("date"), _fmt_amount(e.get("amount", 0)), e.get("voucher_type")]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=col_idx, value=v)
            cell.fill   = _fill(LIGHT_RED)
            cell.border = _border()
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="IN B, NOT IN A").font = _font(bold=True, size=11)
    row += 1
    _apply_header_row(ws, ["Date", "Amount", "Voucher Type"], row=row, bg=RED)
    row += 1
    for e in results_pay.get("_b_not_in_a", []):
        vals = [e.get("date"), _fmt_amount(e.get("amount", 0)), e.get("voucher_type")]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=col_idx, value=v)
            cell.fill   = _fill(LIGHT_RED)
            cell.border = _border()
        row += 1

    _auto_width(ws)


# ─────────────────────────────────────────────
# Sheet 7 — No Invoice Number
# FIX [3]: new sheet for blank invoice entries from both parties
# ─────────────────────────────────────────────

def _build_no_invoice_number(ws, entries: list, party_a: str, party_b: str):
    ws.title = "No Invoice Number"
    headers = ["Party", "Date", "Voucher Type", "Amount", "Reason"]
    _apply_header_row(ws, headers, bg=PURPLE)

    for row_idx, e in enumerate(entries, start=2):
        source = e.get("source", "")
        party  = party_a if source == "A" else party_b if source == "B" else source
        _, date, vtype, amt = _entry_cols(e)
        reason = e.get("unmatched_reason") or "No invoice number found"
        vals = [party, date, vtype, _fmt_amount(amt), reason]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.fill      = _fill(LIGHT_PURPLE)
            cell.alignment = _left()
            cell.border    = _border()

    _auto_width(ws)


# ─────────────────────────────────────────────
# Public: generate_excel
# ─────────────────────────────────────────────

def generate_excel(
    results: dict,
    party_a_name: str,
    party_b_name: str,
    period: dict,
) -> bytes:
    """Build a 7-sheet workbook and return bytes."""
    wb = openpyxl.Workbook()

    inv = results["invoices"]
    pay = results["payments"]

    # Sheet 1 — Summary
    ws1 = wb.active
    _build_summary(ws1, results, party_a_name, party_b_name, period)

    # Sheet 2 — Matched Exact
    ws2 = wb.create_sheet()
    _build_matched_exact(ws2, inv.get("_matched_exact", []), party_a_name, party_b_name)

    # Sheet 3 — Matched Variation
    ws3 = wb.create_sheet()
    _build_matched_variation(ws3, inv.get("_matched_variation", []), party_a_name, party_b_name)

    # Sheet 4 — A Not in B (FIX [1]: standardised name)
    ws4 = wb.create_sheet()
    _build_unmatched_single(ws4, inv.get("_a_not_in_b", []), "A Not in B")

    # Sheet 5 — B Not in A (FIX [1]: standardised name)
    ws5 = wb.create_sheet()
    _build_unmatched_single(ws5, inv.get("_b_not_in_a", []), "B Not in A")

    # Sheet 6 — Payment Mismatches
    ws6 = wb.create_sheet()
    _build_payment_mismatches(ws6, pay, party_a_name, party_b_name)

    # Sheet 7 — No Invoice Number (FIX [3]: new sheet)
    ws7 = wb.create_sheet()
    _build_no_invoice_number(ws7, inv.get("_no_invoice_number", []), party_a_name, party_b_name)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
